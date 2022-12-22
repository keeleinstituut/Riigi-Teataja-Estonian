# reads act URLs file for each act. Writes new .xlsx file of each ET-EN act. Writes all acts metadata in a separate file.
# loeb URLlistist iga akti, kirjutab eraldi .xlsx paralleelsesse ET-EN faili. Kõigi aktide andmetest eraldi metafail.

import openpyxl
from pathlib import Path
import requests
from bs4 import BeautifulSoup
import re
import xlsxwriter
import os
from os.path import exists
from datetime import datetime, date

from FileWriter_functions.domainwriter import subcatanalyzer, subcat_from_main, subjoiner
from FileWriter_functions.getmetainfo_functions import *


def urlcollector(actspath, metapath, filepath):
    wb_obj = openpyxl.load_workbook(actspath) 
    sheet = wb_obj.active
    counter = 0
    # creates metacontent file
    metaworkbook = xlsxwriter.Workbook(metapath) # starts a new .xlsx file
    metaworksheet = metaworkbook.add_worksheet()
    metano = 1
    metaworksheet.write('A1', "filename")
    metaworksheet.write('B1', "id")
    metaworksheet.write('C1', "publishing_year")
    metaworksheet.write('D1', "domain_systematic")
    metaworksheet.write('E1', "domain_Eurovoc")
    metaworksheet.write('F1', "domain_KOV")
    metaworksheet.write('G1', "domain_foregin")
    metaworksheet.write('H1', "issuer")
    metaworksheet.write('I1', "act_type")
    metaworksheet.write('J1', "text_type")
    metaworksheet.write('K1', "in_force_from")
    metaworksheet.write('L1', "in_force_until")
    metaworksheet.write('M1', "validity")
    metaworksheet.write('N1', "publishing_note")
    metaworksheet.write('O1', "title")
    metaworksheet.write('P1', "abbrevation")
    metaworksheet.write('Q1', "act_passed")
    metaworksheet.write('R1', "crawl_date")
    metaworksheet.write('S1', "crawl_time")
    metaworksheet.write('T1', "language")
    metaworksheet.write('U1', "url")
    
    # loeb ridahaaval ja toimetab iga URLiga edasi
    for row in sheet.iter_rows():
        for cell in row:
            url = cell.value
            result = actwriter(url, metaworksheet, metano, filepath)
            metano = result[1]
            counter +=1
            print(counter, "done: ", url, "metadata row:", metano)
    metaworkbook.close() 
    wb_obj.close() 
    print("DONE")

def actwriter(url, metaworksheet, metano, filepath):
    # parse act
    page = requests.get(url)
    soup = BeautifulSoup(page.content, "lxml")

    metacontent = soup.find("table", class_="meta") # metadata
    bodycontent = soup.find("div", id="article-content") # body text

    # finds "vastu võetud" for abbrevation purposes
    vastu_voetud = soup.find("p", class_="vv")
    if vastu_voetud.text != "":
        vv = vastu_voetud.br.previous_sibling.text
        vv = vv.replace('Vastu võetud ','')
    else:
        vv = "None"

    # writes .xlsx file
    fileid = url.split('/')[-1]
    #filename = 'RT-' + str(fileid) + '.xlsx'
    #path = os.path.join(filepath, filename)
    #workbook = xlsxwriter.Workbook(path) # starts a new .xlsx file
    #worksheet = workbook.add_worksheet()
    no = 1

    for br in bodycontent.find_all("br"):
        br.replace_with("\n")
    bodyparagraphs = bodycontent.find_all(["h1", "h2", "h3", "p"]) 
    if not bodyparagraphs[-1].text == "Muudetud järgmiste aktidega (näita)":
        
        filename2 = 'RT-' +str(fileid) + ".xml"
        path = os.path.join(filepath, filename2)
        if not exists(path):
            with open(path, 'w',encoding='utf8') as f:
                
                # writes act metadata
                subcat = "NULL"     
                metadataET = metaparser(metacontent, soup, subcat, url, metaworksheet, fileid, metano, vv, bodyparagraphs)
                subcat = metadataET[0]
                metano = metadataET[2]
                #worksheet.write('A'+str(no),metadataET[1])
                f.write(metadataET[1])
                f.write('\n')

                # writes act body text
                paragraphs = len(bodyparagraphs)-1 # number of paragraphs found
                n = 0 # paragraph index
                while n<=paragraphs:
                    tulemus = paralyzer(bodyparagraphs, n, paragraphs, no)
                    if len(tulemus[1]) != 0:
                        n = tulemus[0]
                        paraname = tulemus[2]
                        paratext = "<"+paraname+">"+tulemus[1]+"</"+paraname+">"
                        no += 1
                        f.write(paratext)
                        f.write('\n')
                        #worksheet.write('A'+str(no),paratext)
                    else:
                        n += 1
                no += 1
                #worksheet.write('A'+str(no),"</doc>")
                f.write("</doc>")

                print("---------------- kirjutas faili ---------------", fileid)
            #workbook.close() 
    
    return fileid, metano


def metaparser(results, soup, subcat, url, metaworksheet, fileid, metano, vv, bodyparagraphs):   
    #filename and file id
    filename = 'filename='+'"RT-'+str(fileid)+'"'        
    idno = 'id='+'"'+str(fileid)+'"'

    #url
    url_meta = str(url) 
    url_actmeta = 'url="'+str(url)+'"' 

    # current date and time
    now = datetime.now() 
    crawltime = now.strftime("%H:%M:%S")
    crawldate = now.strftime("%Y-%m-%d") 
    crawl_date = 'crawl_date="'+crawldate+'"'
    crawl_time = 'crawl_time="'+crawltime+'"'

    # other
    passed = 'act_passed="'+vv+'"'
    lang = 'language="Estonian"'

    # subcategories
    subcat = subcatanalyzer(url) 
    subcat1 = subjoiner(subcat[0])
    subcat2 = subjoiner(subcat[1])
    subcat3 = subjoiner(subcat[2])
    subcat4 = subjoiner(subcat[3])

    if subcat1 == subcat2 == subcat3 == subcat4 == "None":
        for paragraph in bodyparagraphs:        
            if "Määrus kehtestatakse" in paragraph.text:
                subcats = subcat_from_main(paragraph)
                subcat1 = subcats[0]
                subcat2 = subcats[1]
                subcat3 = subcats[2]
                subcat4 = subcats[3]
                break

    sub1 = 'domain_systematic="'+subcat1+'"' 
    sub2 = 'domain_Eurovoc="'+subcat2+'"' 
    sub3 = 'domain_KOV="'+subcat3+'"' 
    sub4 = 'domain_foregin="'+subcat4+'"' 

    # other metainfo
    metainfo1 = results.find_all("th") # key             
    metainfo2 = results.find_all("td") # value
    metainfo1 = [x.text.replace(":","") for x in metainfo1]
    metainfo2 = [" ".join(x.text.split()) for x in metainfo2]
    res_meta = {metainfo1[i]: metainfo2[i] for i in range(len(metainfo1))} # dictionary
    
    titleabb_res = get_titleabb(soup)
    title_long = titleabb_res[0]
    title = titleabb_res[1]
    abbrevation = titleabb_res[2]
    abbrevation_act = titleabb_res[3]

    issuer_res = get_issuer(res_meta)
    issuer = issuer_res[0]
    metaissuer = issuer_res[1]
    
    act_tye_res = get_acttype(res_meta, metaissuer)
    act_type = act_tye_res[0]
    metaact_type = act_tye_res[1] 
    
    text_type_res = get_texttype(res_meta)
    text_type = text_type_res[0]
    metatext_type = text_type_res[1]
    
    in_force_from_res = get_inforcefrom(res_meta, vv)
    in_force_from = in_force_from_res[0]
    metain_force_from = in_force_from_res[1]

    in_force_until_res = get_inforceuntil(res_meta)
    in_force_until = in_force_until_res[0]
    metain_force_until = in_force_until_res[1]
    
    publishing_note_res = get_publishingnote(res_meta)
    publishing_note = publishing_note_res[0]
    metapublishing_note = publishing_note_res[1]

    validity_res = get_validity(metain_force_until, now)
    validity_note = validity_res[0]
    metavalidity_note = validity_res[1]
   
    timestamp_res = get_timestamp(metain_force_from, metapublishing_note)
    timestamp_yearmonthday = timestamp_res[0]
    publishing_year = timestamp_res[1]   
    
    # writes metainfo to metafile
    metano += 1
    metaworksheet.write('A'+str(metano), "RT-"+str(fileid))
    metaworksheet.write('B'+str(metano), fileid)
    metaworksheet.write('C'+str(metano), timestamp_yearmonthday)
    metaworksheet.write('D'+str(metano), subcat1)
    metaworksheet.write('E'+str(metano), subcat2)
    metaworksheet.write('F'+str(metano), subcat3)
    metaworksheet.write('G'+str(metano), subcat4)
    metaworksheet.write('H'+str(metano), metaissuer)
    metaworksheet.write('I'+str(metano), metaact_type)
    metaworksheet.write('J'+str(metano), metatext_type)
    metaworksheet.write('K'+str(metano), metain_force_from)
    metaworksheet.write('L'+str(metano), metain_force_until)
    metaworksheet.write('M'+str(metano), metavalidity_note)
    metaworksheet.write('N'+str(metano), metapublishing_note)
    metaworksheet.write('O'+str(metano), title_long)
    metaworksheet.write('P'+str(metano), abbrevation)
    metaworksheet.write('Q'+str(metano), vv)
    metaworksheet.write('R'+str(metano), crawldate)
    metaworksheet.write('S'+str(metano), crawltime)
    metaworksheet.write('T'+str(metano), 'Estonian')
    metaworksheet.write('U'+str(metano), url_meta)
    
    # creates meta for act file
    metadata = " ".join(['<doc', filename, idno, publishing_year,  sub1, sub2, sub3,  sub4, issuer, act_type, text_type,  in_force_from, in_force_until, validity_note, publishing_note, title, abbrevation_act, passed, crawl_date, crawl_time, lang, url_actmeta,'>'])

    print("kirjutas metafaili: ", fileid)
    return subcat, metadata, metano



def paralyzer(bodyparagraphs, n, paragraphs, no):
    value = ""
    paraname = bodyparagraphs[n].name
    para1 = bodyparagraphs[n].text
    para1 = ' '.join((' '.join(para1.splitlines())).split())
    n += 1
    value += para1

    return n, value, paraname


